import csv
import io
from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload

from ..auth import create_access_token, get_current_admin, verify_password
from ..database import get_db
from ..document_utils import validate_document_upload
from ..models import AdminUser, AuditLog, Child, Document, GrandChild, Registration
from ..rate_limit import rate_limit
from ..schemas import (
    AuditLogOut,
    DocumentOut,
    ImportRowError,
    ImportSummaryOut,
    LoginIn,
    MarkSeenIn,
    RegistrationIn,
    RegistrationListOut,
    RegistrationOut,
    StatusUpdateIn,
    TokenOut,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])

REGISTRATION_COLUMNS = [
    ("id", "ID"),
    ("created_at", "Submitted At"),
    ("status", "Status"),
    ("kgoro", "Kgoro"),
    ("mokgomane", "Mokgomane"),
    ("section", "Section"),
    ("receipt_number", "Receipt Number"),
    ("reference_no", "Reference No"),
    ("stand_no", "Stand No"),
    ("zone", "Zone"),
    ("original_member_title", "Main Member Title (Orig. Dispossessed)"),
    ("original_member_name", "Main Member Name (Orig. Dispossessed)"),
    ("original_member_id_number", "Main Member ID (Orig. Dispossessed)"),
    ("original_spouse_title", "Spouse Title (Orig. Dispossessed)"),
    ("original_spouse_name", "Spouse Name (Orig. Dispossessed)"),
    ("original_spouse_id_number", "Spouse ID (Orig. Dispossessed)"),
    ("claimant_title", "Main Member Title (Main Claimant)"),
    ("claimant_name", "Main Member Name (Main Claimant)"),
    ("claimant_id_number", "Main Member ID (Main Claimant)"),
    ("claimant_spouse_title", "Spouse Title (Main Claimant)"),
    ("claimant_spouse_name", "Spouse Name (Main Claimant)"),
    ("claimant_spouse_id_number", "Spouse ID (Main Claimant)"),
    ("relationship_to_odi", "Relationship to Odi"),
    ("email", "Email"),
    ("place_of_origin", "Place of Origin / Ancestral Place of Origin"),
    ("business_details", "Business Details / Additional Details"),
    ("origin_ethnicity", "Origin / Ethnicity"),
    ("family_representative", "Family Representative"),
    ("power_of_attorney", "Power of Attorney"),
]

# Columns that are safe to bulk-import/update from a CSV. Excludes id/created_at
# (id is only used to detect an update-vs-create) and status/review fields
# (reviewed separately through the approve/reject workflow, not spreadsheet edits).
IMPORTABLE_FIELDS = [field for field, _ in REGISTRATION_COLUMNS if field not in ("id", "created_at", "status")]


@router.post(
    "/login",
    response_model=TokenOut,
    dependencies=[Depends(rate_limit(max_requests=10, window_seconds=900))],
)
def login(payload: LoginIn, db: Session = Depends(get_db)):
    admin = db.query(AdminUser).filter(AdminUser.username == payload.username).first()
    if not admin or not verify_password(payload.password, admin.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password",
        )
    token = create_access_token(subject=admin.username)
    return TokenOut(access_token=token)


def _registration_label(reg: Registration) -> str:
    title = reg.claimant_title or reg.original_member_title or ""
    name = reg.claimant_name or reg.original_member_name or "Unknown"
    return f"#{reg.id} {title} {name}".strip()


def _log_action(
    db: Session,
    admin_username: str,
    action: str,
    registration_id: int | None = None,
    registration_label: str | None = None,
    detail: str | None = None,
):
    db.add(
        AuditLog(
            registration_id=registration_id,
            registration_label=registration_label,
            admin_username=admin_username,
            action=action,
            detail=detail,
        )
    )


def _load_query(db: Session):
    return db.query(Registration).options(
        joinedload(Registration.children),
        joinedload(Registration.grandchildren),
        joinedload(Registration.documents),
    )


def _annotate_is_new(registrations: list[Registration]) -> list[Registration]:
    for reg in registrations:
        reg.is_new = reg.status == "pending" and not reg.seen_by_admin
    return registrations


def _apply_filters(
    query,
    status_filter=None,
    kgoro=None,
    section=None,
    zone=None,
    place_of_origin=None,
    origin_ethnicity=None,
    q=None,
    date_from=None,
    date_to=None,
):
    if status_filter:
        query = query.filter(Registration.status == status_filter)
    if kgoro:
        query = query.filter(Registration.kgoro.ilike(f"%{kgoro}%"))
    if section:
        query = query.filter(Registration.section.ilike(f"%{section}%"))
    if zone:
        query = query.filter(Registration.zone.ilike(f"%{zone}%"))
    if place_of_origin:
        query = query.filter(Registration.place_of_origin.ilike(f"%{place_of_origin}%"))
    if origin_ethnicity:
        query = query.filter(Registration.origin_ethnicity.ilike(f"%{origin_ethnicity}%"))
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Registration.claimant_name.ilike(like),
                Registration.original_member_name.ilike(like),
                Registration.claimant_id_number.ilike(like),
                Registration.original_member_id_number.ilike(like),
            )
        )
    if date_from:
        query = query.filter(Registration.created_at >= date_from)
    if date_to:
        query = query.filter(Registration.created_at <= date_to)
    return query


class RegistrationFilters:
    def __init__(
        self,
        status: str | None = Query(None),
        kgoro: str | None = Query(None),
        section: str | None = Query(None),
        zone: str | None = Query(None),
        place_of_origin: str | None = Query(None),
        origin_ethnicity: str | None = Query(None),
        q: str | None = Query(None),
        date_from: datetime | None = Query(None),
        date_to: datetime | None = Query(None),
    ):
        self.status = status
        self.kgoro = kgoro
        self.section = section
        self.zone = zone
        self.place_of_origin = place_of_origin
        self.origin_ethnicity = origin_ethnicity
        self.q = q
        self.date_from = date_from
        self.date_to = date_to

    def apply(self, query):
        return _apply_filters(
            query,
            status_filter=self.status,
            kgoro=self.kgoro,
            section=self.section,
            zone=self.zone,
            place_of_origin=self.place_of_origin,
            origin_ethnicity=self.origin_ethnicity,
            q=self.q,
            date_from=self.date_from,
            date_to=self.date_to,
        )


@router.get("/registrations", response_model=RegistrationListOut)
def list_registrations(
    filters: RegistrationFilters = Depends(),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
    db: Session = Depends(get_db),
    _admin: AdminUser = Depends(get_current_admin),
):
    base_query = filters.apply(db.query(Registration))
    total = base_query.count()
    pending_count = base_query.filter(Registration.status == "pending").count()
    new_count = base_query.filter(
        Registration.status == "pending", Registration.seen_by_admin.is_(False)
    ).count()
    children_count = (
        filters.apply(db.query(Child).join(Registration, Child.registration_id == Registration.id)).count()
    )
    grandchildren_count = (
        filters.apply(
            db.query(GrandChild).join(Registration, GrandChild.registration_id == Registration.id)
        ).count()
    )

    items_query = filters.apply(_load_query(db))
    registrations = (
        items_query.order_by(Registration.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    _annotate_is_new(registrations)

    return RegistrationListOut(
        items=[RegistrationOut.model_validate(r, from_attributes=True) for r in registrations],
        total=total,
        page=page,
        page_size=page_size,
        pending_count=pending_count,
        new_count=new_count,
        children_count=children_count,
        grandchildren_count=grandchildren_count,
    )


def _build_workbook(registrations: list[Registration]) -> Workbook:
    wb = Workbook()

    main_sheet = wb.active
    main_sheet.title = "Registrations"
    main_sheet.append([label for _, label in REGISTRATION_COLUMNS])
    for reg in registrations:
        row = []
        for field, _ in REGISTRATION_COLUMNS:
            value = getattr(reg, field)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            row.append(value)
        main_sheet.append(row)

    children_sheet = wb.create_sheet("Children")
    children_sheet.append(["Registration ID", "Main Member Name", "Child Name", "ID Number", "Gender", "Contact"])
    for reg in registrations:
        for child in reg.children:
            children_sheet.append(
                [reg.id, reg.claimant_name or reg.original_member_name, child.name, child.id_number, child.gender, child.contact]
            )

    grandchildren_sheet = wb.create_sheet("Grandchildren")
    grandchildren_sheet.append(["Registration ID", "Main Member Name", "Grandchild Name", "ID Number", "Gender"])
    for reg in registrations:
        for gc in reg.grandchildren:
            grandchildren_sheet.append(
                [reg.id, reg.claimant_name or reg.original_member_name, gc.name, gc.id_number, gc.gender]
            )

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 45)

    return wb


@router.get("/registrations/export")
def export_registrations(
    filters: RegistrationFilters = Depends(),
    db: Session = Depends(get_db),
    _admin: AdminUser = Depends(get_current_admin),
):
    query = filters.apply(
        db.query(Registration).options(joinedload(Registration.children), joinedload(Registration.grandchildren))
    )
    registrations = query.order_by(Registration.created_at.asc()).all()

    wb = _build_workbook(registrations)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"bbmt_membership_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/registrations/export.csv")
def export_registrations_csv(
    filters: RegistrationFilters = Depends(),
    db: Session = Depends(get_db),
    _admin: AdminUser = Depends(get_current_admin),
):
    query = filters.apply(db.query(Registration))
    registrations = query.order_by(Registration.created_at.asc()).all()

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in REGISTRATION_COLUMNS])
    for reg in registrations:
        row = []
        for field, _ in REGISTRATION_COLUMNS:
            value = getattr(reg, field)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            row.append(value)
        writer.writerow(row)

    filename = f"bbmt_membership_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/registrations/import", response_model=ImportSummaryOut)
async def import_registrations_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    raw = await file.read()
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    importable_set = set(IMPORTABLE_FIELDS)
    label_to_field = {label: field for field, label in REGISTRATION_COLUMNS if field in importable_set}
    known_headers = set(label_to_field) | importable_set | {"id", "ID"}

    created = 0
    updated = 0
    skipped = 0
    errors: list[ImportRowError] = []

    for row_number, row in enumerate(reader, start=2):  # header is row 1
        try:
            mapped = {}
            for header, value in row.items():
                if header is None:
                    continue
                field = label_to_field.get(header, header if header in IMPORTABLE_FIELDS else None)
                if field:
                    mapped[field] = value.strip() if isinstance(value, str) else value
                elif header not in known_headers and header.strip():
                    pass  # ignore unrecognized columns rather than failing the row

            raw_id = row.get("ID") or row.get("id")
            registration = None
            if raw_id and str(raw_id).strip():
                try:
                    registration = db.query(Registration).filter(Registration.id == int(raw_id)).first()
                except ValueError:
                    registration = None

            if registration:
                for field, value in mapped.items():
                    setattr(registration, field, value or None)
                updated += 1
            else:
                registration = Registration(**{k: (v or None) for k, v in mapped.items()})
                db.add(registration)
                created += 1
        except Exception as exc:  # noqa: BLE001 - collect per-row errors, don't abort the batch
            skipped += 1
            errors.append(ImportRowError(row=row_number, error=str(exc)))

    _log_action(
        db,
        admin.username,
        "created_via_import",
        None,
        f"CSV import: {file.filename or 'upload'}",
        detail=f"created={created} updated={updated} skipped={skipped}",
    )
    db.commit()
    return ImportSummaryOut(created=created, updated=updated, skipped=skipped, errors=errors[:50])


@router.post("/registrations/mark-seen")
def mark_registrations_seen(
    payload: MarkSeenIn,
    db: Session = Depends(get_db),
    _admin: AdminUser = Depends(get_current_admin),
):
    query = db.query(Registration)
    if payload.ids:
        query = query.filter(Registration.id.in_(payload.ids))
    else:
        query = query.filter(Registration.status == "pending")
    query.update({Registration.seen_by_admin: True}, synchronize_session=False)
    db.commit()
    return {"ok": True}


@router.get("/registrations/{registration_id}", response_model=RegistrationOut)
def get_registration(
    registration_id: int,
    db: Session = Depends(get_db),
    _admin: AdminUser = Depends(get_current_admin),
):
    registration = _load_query(db).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    registration.is_new = registration.status == "pending" and not registration.seen_by_admin
    return registration


@router.patch("/registrations/{registration_id}", response_model=RegistrationOut)
def update_registration(
    registration_id: int,
    payload: RegistrationIn,
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    data = payload.model_dump(exclude={"children", "grandchildren"})
    for field, value in data.items():
        setattr(registration, field, value)

    registration.children.clear()
    for child in payload.children:
        if any(getattr(child, field) for field in child.model_fields):
            registration.children.append(Child(**child.model_dump()))

    registration.grandchildren.clear()
    for grandchild in payload.grandchildren:
        if any(getattr(grandchild, field) for field in grandchild.model_fields):
            registration.grandchildren.append(GrandChild(**grandchild.model_dump()))

    _log_action(db, admin.username, "edited", registration.id, _registration_label(registration))

    db.commit()
    db.refresh(registration)
    registration.is_new = registration.status == "pending" and not registration.seen_by_admin
    return registration


@router.patch("/registrations/{registration_id}/status", response_model=RegistrationOut)
def update_registration_status(
    registration_id: int,
    payload: StatusUpdateIn,
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    registration.status = payload.status
    registration.review_note = payload.review_note
    registration.reviewed_at = datetime.utcnow()
    registration.reviewed_by = admin.username

    _log_action(
        db,
        admin.username,
        payload.status,
        registration.id,
        _registration_label(registration),
        detail=payload.review_note,
    )

    db.commit()
    db.refresh(registration)
    registration.is_new = registration.status == "pending" and not registration.seen_by_admin
    return registration


@router.delete("/registrations/{registration_id}", status_code=204)
def delete_registration(
    registration_id: int,
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    _log_action(db, admin.username, "deleted", registration.id, _registration_label(registration))

    db.delete(registration)
    db.commit()


@router.post("/registrations/{registration_id}/documents", response_model=DocumentOut, status_code=201)
async def upload_admin_document(
    registration_id: int,
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    registration = db.query(Registration).filter(Registration.id == registration_id).first()
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")

    contents = await file.read()
    validate_document_upload(doc_type, file, contents)

    doc = Document(
        registration_id=registration_id,
        doc_type=doc_type,
        filename=file.filename,
        content_type=file.content_type,
        data=contents,
    )
    db.add(doc)
    _log_action(
        db, admin.username, "document_uploaded", registration.id, _registration_label(registration), detail=doc_type
    )
    db.commit()
    db.refresh(doc)
    return doc


@router.get("/documents/{document_id}")
def get_document(
    document_id: int,
    db: Session = Depends(get_db),
    _admin: AdminUser = Depends(get_current_admin),
):
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return StreamingResponse(
        io.BytesIO(doc.data),
        media_type=doc.content_type,
        headers={"Content-Disposition": f'inline; filename="{doc.filename or "document.png"}"'},
    )


@router.delete("/documents/{document_id}", status_code=204)
def delete_document(
    document_id: int,
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    doc = db.query(Document).filter(Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    registration = db.query(Registration).filter(Registration.id == doc.registration_id).first()
    _log_action(
        db,
        admin.username,
        "document_deleted",
        doc.registration_id,
        _registration_label(registration) if registration else None,
        detail=doc.doc_type,
    )

    db.delete(doc)
    db.commit()


@router.get("/audit-log", response_model=list[AuditLogOut])
def list_audit_log(
    registration_id: int | None = Query(None),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    _admin: AdminUser = Depends(get_current_admin),
):
    query = db.query(AuditLog)
    if registration_id is not None:
        query = query.filter(AuditLog.registration_id == registration_id)
    return query.order_by(AuditLog.created_at.desc()).limit(limit).all()
